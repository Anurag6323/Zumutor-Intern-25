from Bio import SeqIO
import re
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# === Define CDR ranges (1-based, inclusive) ===
CDR_L1_RANGE = (24, 38)
CDR_H2_RANGE = (50, 68)
CDR_H3_RANGE = (101, 110)
# ==============================================

cdr_l1_range = (CDR_L1_RANGE[0] - 1, CDR_L1_RANGE[1])
cdr_h2_range = (CDR_H2_RANGE[0] - 1, CDR_H2_RANGE[1])
cdr_h3_range = (CDR_H3_RANGE[0] - 1, CDR_H3_RANGE[1])

input_fasta = "trem1_variants.fasta"
output_excel = "trem1_cdrs_per_residue.xlsx"

mutations = defaultdict(dict)

# Parse FASTA and group by mutation ID
for record in SeqIO.parse(input_fasta, "fasta"):
    header = record.description
    sequence = str(record.seq)

    match = re.match(r"(mut\d+)_([LH])\|([^|]*)\|", header)
    if not match:
        print(f"Warning: Skipping unrecognized header format: {header}")
        continue

    mut_id, chain_type, mut_list_str = match.groups()
    mut_list = [m.strip() for m in mut_list_str.split(",") if m.strip()]
    mutations[mut_id][chain_type] = {
        "seq": sequence,
        "mutations": mut_list
    }

# Extract mutated residue positions (0-based index)
def get_mutated_positions(mut_list, cdr_range):
    positions = set()
    for mut in mut_list:
        match = re.match(r"([A-Z])(\d+)([A-Z])", mut)
        if match:
            wt, pos_str, new = match.groups()
            pos = int(pos_str) - 1
            if cdr_range[0] <= pos < cdr_range[1]:
                positions.add(pos - cdr_range[0])
    return positions

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "CDR_Residues"

# Define color style
fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
bold = Font(bold=True)

# Write header
cdr_lengths = {
    "CDR_L1": cdr_l1_range[1] - cdr_l1_range[0],
    "CDR_H2": cdr_h2_range[1] - cdr_h2_range[0],
    "CDR_H3": cdr_h3_range[1] - cdr_h3_range[0]
}
header = ["Mutation_ID"]
for name, length in cdr_lengths.items():
    header += [f"{name}_{i+1}" for i in range(length)]
ws.append(header)

# Write data
for mut_id, chains in mutations.items():
    if "L" not in chains or "H" not in chains:
        print(f"Skipping {mut_id} due to missing chain")
        continue

    light_seq = chains["L"]["seq"]
    light_muts = chains["L"]["mutations"]
    heavy_seq = chains["H"]["seq"]
    heavy_muts = chains["H"]["mutations"]

    # Get CDRs
    cdr_l1_seq = light_seq[cdr_l1_range[0]:cdr_l1_range[1]]
    cdr_h2_seq = heavy_seq[cdr_h2_range[0]:cdr_h2_range[1]]
    cdr_h3_seq = heavy_seq[cdr_h3_range[0]:cdr_h3_range[1]]

    # Get positions of mutations in each CDR
    mut_l1_pos = get_mutated_positions(light_muts, cdr_l1_range)
    mut_h2_pos = get_mutated_positions(heavy_muts, cdr_h2_range)
    mut_h3_pos = get_mutated_positions(heavy_muts, cdr_h3_range)

    # Build row
    row = [mut_id]
    row += list(cdr_l1_seq)
    row += list(cdr_h2_seq)
    row += list(cdr_h3_seq)
    ws.append(row)

    # Highlight mutated residues
    row_num = ws.max_row
    col_offset = 2  # Mutation_ID is column 1

    # Apply formatting
    for i, pos in enumerate(mut_l1_pos):
        cell = ws.cell(row=row_num, column=col_offset + pos)
        cell.fill = fill
        cell.font = bold

    col_offset += cdr_lengths["CDR_L1"]
    for i, pos in enumerate(mut_h2_pos):
        cell = ws.cell(row=row_num, column=col_offset + pos)
        cell.fill = fill
        cell.font = bold

    col_offset += cdr_lengths["CDR_H2"]
    for i, pos in enumerate(mut_h3_pos):
        cell = ws.cell(row=row_num, column=col_offset + pos)
        cell.fill = fill
        cell.font = bold

# Save workbook
wb.save(output_excel)
print(f"\n✅ Done! Excel file with per-residue highlights saved to '{output_excel}'")
